from openpyxl import load_workbook


wb = load_workbook("c:/Users/Martin/Documents/test.xlsx")
ws1 = wb['Tabelle1']
max_row=ws1.max_row
max_column=ws1.max_column
print(max_row)
print(max_column)


values = [["Martin", "Mitel", "Innovaphone", "Sirio"],["Ferdi", "Innovaphone", "Sirio"], ["Andrea", "Sirio"]]
xvalues = ["Martin", "Ferdi"]
yvalues = ["Mitel", "Innovaphone", "Sirio"]

for c in range(1,max_column+1):
    for v in range(len(values)):
        if ws1.cell(row=1, column=c).value in values[v]:
            #print(ws1.cell(row=1, column=i).value)
            for r in range(1,max_row+1):
                if ws1.cell(row=r, column=1).value in values[v]:
                    ws1.cell(row=r, column=c).value = "x"

wb.save("c:/Users/Martin/Documents/test2.xlsx")







