from openpyxl import load_workbook
from datetime import datetime

class Exportmxone():
    def __init__(self, exportpath, customerdata, userdata, systemdata, externalnumers):
        self.exportpath = exportpath
        self.customerdata = customerdata
        self.userdata = userdata
        self.systemdata = systemdata
        self.externalnumbers = externalnumers

    def writemxonedata(self):
        # Start Row für Userdaten in der Excel Row Userdata
        self.ru = 8
        # Start Row für Durchwahlen Row Durchwahlen
        self.rd = 10
        # Start Row für Systemdaten
        self.rs = 7

        # Pfad der Vorlage
        self.wbmx = load_workbook("./files/PBX_Doku_Vorlage.xlsx")
        # Sheet auswählen wo die Daten hingeschrieben werden
        self.wsmxuser = self.wbmx['Kundendatenliste']
        # Sheet auswähen wo die Durchwahlen hingerschrieben werden
        self.wsmxdw = self.wbmx['Durchwahl Liste']
        # Sheet auswähen wo die Systemdaten hingerschrieben werden
        self.wsmxsystem = self.wbmx['System']

        # Kundenname
        self.wsmxuser.cell(row=2, column=2).value = self.customerdata["Kundenname"]
        # Datum
        self.wsmxuser.cell(row=2, column=8).value = datetime.now().strftime("%d.%m.%Y")
        # Adresse
        self.wsmxuser.cell(row=3, column=2).value = self.customerdata["Adresse"]
        # Ort
        self.wsmxuser.cell(row=4, column=2).value = self.customerdata["Ort"]
        # Hauptnummer
        self.wsmxuser.cell(row=2, column=5).value = self.customerdata["Hauptnummer"]
        # Faxnummer
        self.wsmxuser.cell(row=3, column=5).value = self.customerdata["Faxnummer"]
        # Nummerbereich 1
        self.wsmxuser.cell(row=2, column=11).value = str(self.customerdata["Nummerbereichfrom1"]) + " - " + str(self.customerdata["Nummerbereichto1"])
        # Nummerbereich 2
        self.wsmxuser.cell(row=3, column=11).value = str(self.customerdata["Nummerbereichfrom2"]) + " - " + str(self.customerdata["Nummerbereichto2"])
        # Nummerbereich 3
        self.wsmxuser.cell(row=4, column=11).value = str(self.customerdata["Nummerbereichfrom3"]) + " - " + str(self.customerdata["Nummerbereichto3"])

        # Schreibt alle Systemdaten in das Excel
        # MX-One Version
        self.wsmxsystem.cell(row=3, column=2).value = str(self.systemdata["mxversion"])
        # SNM Version
        self.wsmxsystem.cell(row=4, column=2).value = str(self.systemdata["snmversion"])
        # PM Version
        self.wsmxsystem.cell(row=5, column=2).value = str(self.systemdata["pmversion"])

        for i in self.systemdata:
            if "Cassandra" in i:
                self.wsmxsystem.cell(row=self.rs, column=1).value = str(i)
                self.wsmxsystem.cell(row=self.rs, column=2).value = "Lim IP: " + str(self.systemdata[i][0])
                self.wsmxsystem.cell(row=self.rs, column=3).value = "DB IP: " + str(self.systemdata[i][1])
                self.rs = self.rs + 1
            elif "lim" in i:
                self.wsmxsystem.cell(row=self.rs, column=1).value = str(i)
                self.wsmxsystem.cell(row=self.rs, column=2).value = str(self.systemdata[i][0])
                self.wsmxsystem.cell(row=self.rs, column=3).value = "IP: " + str(self.systemdata[i][1])
                self.rs = self.rs + 1
            elif "gateway" in i:
                self.wsmxsystem.cell(row=self.rs, column=1).value = str(i)
                self.wsmxsystem.cell(row=self.rs, column=2).value = "MGU: " + str(self.systemdata[i][0])
                self.wsmxsystem.cell(row=self.rs, column=3).value = "SW: " + str(self.systemdata[i][1])
                self.wsmxsystem.cell(row=self.rs, column=4).value = "Typ: " + str(self.systemdata[i][2])
                self.wsmxsystem.cell(row=self.rs, column=5).value = "HW Rev: " + str(self.systemdata[i][3])
                self.wsmxsystem.cell(row=self.rs, column=6).value = "SN: " + str(self.systemdata[i][4])
                self.rs = self.rs + 1



        # Befüllt das zweite Sheet mit allen Möglichen Durchwahlen
        for i in self.externalnumbers:
            self.wsmxdw.cell(row=self.rd, column=1).value = str(i).replace("41", "0", 1)
            self.rd = self.rd + 1


        # Schreibt die Userdaten in die Excel Datei
        for i in self.userdata:
            # Number
            self.wsmxuser.cell(row=self.ru, column=2).value = self.userdata[i][0]
            # TYP
            self.wsmxuser.cell(row=self.ru, column=3).value = self.userdata[i][1]
            # HW Adresse / IP
            self.wsmxuser.cell(row=self.ru, column=4).value = self.userdata[i][2]
            # Name 1
            self.wsmxuser.cell(row=self.ru, column=7).value = self.userdata[i][3]
            # Name 2
            self.wsmxuser.cell(row=self.ru, column=8).value = self.userdata[i][4]
            # Berechtigungen CSP CAT
            self.wsmxuser.cell(row=self.ru, column=10).value = self.userdata[i][5]
            # Server
            # self.wsmxuser.cell(row=self.ru, column=5).value = self.userdata[i][6]
            # Secondary Dir 1
            self.wsmxuser.cell(row=self.ru, column=12).value = self.userdata[i][7]
            # Secondary Dir 2
            self.wsmxuser.cell(row=self.ru, column=13).value = self.userdata[i][8]
            # Third
            self.wsmxuser.cell(row=self.ru, column=11).value = self.userdata[i][9]
            # Externnummer von Extern nach Intern
            self.wsmxuser.cell(row=self.ru, column=1).value = self.userdata[i][10]
            # CLIP
            self.wsmxuser.cell(row=self.ru, column=9).value = self.userdata[i][11]

            # Row Plus 1
            self.ru = self.ru + 1


        #Speichert die Datei als Kopie
        self.wbmx.save(self.exportpath + "/PBX_Doku_" + self.customerdata["Kundenname"] +
                       datetime.now().strftime("%Y%m%d_%I%M%S") + ".xlsx")


class Exportsirio():
    def __init__(self, exportpath, customerdata, querydata, contdata, espadata, jobdata):
        self.exportpath = exportpath
        self.customerdata = customerdata
        self.querydata = querydata
        self.contdata = contdata
        self.espadata = espadata
        self.jobdata = jobdata

    def writesiriodata(self):
        # Pfad der Vorlage
        self.wbsirio = load_workbook("./files/Sirio_Doku_Vorlage.xlsx")
        # Sheet auswählen wo die Daten hingeschrieben werden
        self.wssirio = self.wbsirio['Alarmmatrix']

        # Start Row für Ereignisse / Alarme
        self.ra = 9
        # Start Column für Jobs / Teilnehmer
        self.cj = 11

        # Kundenname
        self.wssirio.cell(row=2, column=4).value = self.customerdata["Kundenname"]
        # Datum
        self.wssirio.cell(row=4, column=10).value = datetime.now().strftime("%d.%m.%Y")
        # Adresse
        self.wssirio.cell(row=3, column=4).value = self.customerdata["Adresse"]
        # Ort
        self.wssirio.cell(row=4, column=4).value = self.customerdata["Ort"]

        # Alarme / Ereignisse Befüllen
        # Daten Sortieren
        # Eingangskontakte
        self.contdata.sort(key=lambda cont: cont[1])
        self.contdata.sort(key=lambda cont: cont[0])
        # ESPA Alarme
        self.espadata.sort(key=lambda psa: psa[2])

        for i in range(len(self.contdata)):
            self.wssirio.cell(row=self.ra, column=1).value = str(self.contdata[i][0])
            self.wssirio.cell(row=self.ra, column=4).value = str(self.contdata[i][1])
            self.wssirio.cell(row=self.ra, column=9).value = str(self.contdata[i][7])
            if "-" in str(self.contdata[i][2]):
                self.wssirio.cell(row=self.ra, column=6).value = "NC"
            else:
                self.wssirio.cell(row=self.ra, column=6).value = "NO"
            self.ra = self.ra + 1

        for i in range(len(self.espadata)):
            self.wssirio.cell(row=self.ra, column=1).value = "ESPA Incoming"
            self.wssirio.cell(row=self.ra, column=5).value = str(self.espadata[i][2])
            self.wssirio.cell(row=self.ra, column=9).value = str(self.espadata[i][7])
            self.wssirio.cell(row=self.ra, column=7).value = str(self.espadata[i][3])
            self.ra = self.ra + 1

        # Teilnehmer / Jobs Befüllen
        self.jobdata.sort(key=lambda name: name[1])
        self.jobdata.sort(key=lambda nr: nr[4])
        self.jobdata.sort(key=lambda nr: nr[3])

        for i in range(len(self.jobdata)):
            self.wssirio.cell(row=7, column=self.cj).value = str(self.jobdata[i][1])
            self.wssirio.cell(row=6, column=self.cj).value = str(self.jobdata[i][4])
            self.cj = self.cj + 1




        # Query für Kreuze Matrix
        self.max_row = self.wssirio.max_row
        self.max_column = self.wssirio.max_column

        self.ra = 9
        # Start Column für Jobs / Teilnehmer
        self.cj = 11

        for i in self.querydata:
            print(i)
            print(self.querydata[i])

        for c in range(self.cj, self.max_column + 1):
            for q in self.querydata:
                if str(self.wssirio.cell(row=7, column=c).value) in str(self.querydata[q]):
                    for r in range(self.ra, self.max_row + 1):
                        print(q)
                        print(self.wssirio.cell(row=r, column=9).value)
                        if str(self.wssirio.cell(row=r, column=9).value) in str(q):
                            self.wssirio.cell(row=r, column=c).value = "x"


        # Speichert die Datei als Kopie
        self.wbsirio.save(self.exportpath + "/Sirio_Doku_" + self.customerdata["Kundenname"] +
                        datetime.now().strftime("%Y%m%d_%I%M%S") + ".xlsx")

