from openpyxl import load_workbook

from datetime import datetime


class Exportsirio():
    def __init__(self):
        wb = load_workbook("c:/Users/Martin/Documents/test.xlsx")
        ws1 = wb['Tabelle1']
        max_row = ws1.max_row
        max_column = ws1.max_column
        print(max_row)
        print(max_column)

        values = [["Martin", "Mitel", "Innovaphone", "Sirio"], ["Ferdi", "Innovaphone", "Sirio"], ["Andrea", "Sirio"]]
        xvalues = ["Martin", "Ferdi"]
        yvalues = ["Mitel", "Innovaphone", "Sirio"]

        for c in range(1, max_column + 1):
            for v in range(len(values)):
                if ws1.cell(row=1, column=c).value in values[v]:
                    # print(ws1.cell(row=1, column=i).value)
                    for r in range(1, max_row + 1):
                        if ws1.cell(row=r, column=1).value in values[v]:
                            ws1.cell(row=r, column=c).value = "x"

        wb.save("c:/Users/Martin/Documents/test2.xlsx")



class Exportmxone():
    def __init__(self, exportpath, customerdata, userdata, systemdata):
        self.exportpath = exportpath
        self.customerdata = customerdata
        self.userdata = userdata
        self.systemdata = systemdata

    def writemxonedata(self):
        self.r = 8
        self.wbmx = load_workbook("c:/PBX_Doku_Vorlage.xlsx")
        self.wsmxuser = self.wbmx['Kundendatenliste']

        #Kundenname
        self.wsmxuser.cell(row=2, column=2).value = self.customerdata["Kundenname"]
        #Datum
        self.wsmxuser.cell(row=2, column=8).value = datetime.now().strftime("%d.%m.%Y")
        #Adresse
        self.wsmxuser.cell(row=3, column=2).value = self.customerdata["Adresse"]
        #Ort
        self.wsmxuser.cell(row=4, column=2).value = self.customerdata["Ort"]

        #Schreibt die Daten in die Excel Datei
        for i in self.userdata:
            #Number
            self.wsmxuser.cell(row=self.r, column=2).value = self.userdata[i][0]
            #TYP
            self.wsmxuser.cell(row=self.r, column=3).value = self.userdata[i][1]
            self.wsmxuser.cell(row=self.r, column=4).value = self.userdata[i][2]
            #Name 1
            self.wsmxuser.cell(row=self.r, column=7).value = self.userdata[i][3]
            #Name 2
            self.wsmxuser.cell(row=self.r, column=8).value = self.userdata[i][4]
            self.wsmxuser.cell(row=self.r, column=10).value = self.userdata[i][5]
            #Server
            #self.wsmxuser.cell(row=self.r, column=5).value = self.userdata[i][6]
            self.wsmxuser.cell(row=self.r, column=12).value = self.userdata[i][7]
            self.wsmxuser.cell(row=self.r, column=13).value = self.userdata[i][8]
            #Third
            #self.wsmxuser.cell(row=self.r, column=6).value = self.userdata[i][9]
            self.r = self.r + 1

        self.wbmx.save("C:/Users/martinachermann/Desktop/PBX_Doku_" + self.customerdata["Kundenname"] +
                       datetime.now().strftime("%Y%m%d_%I%M%S") + ".xlsx")


        print(self.customerdata)
        print(self.userdata)

    def printmxonedata(self):
        self.r = 8
        for i in self.userdata:
            print(self.userdata[i][3])
            print(self.userdata[i][1])
            print(self.userdata[i])
            self.r = self.r + 1












